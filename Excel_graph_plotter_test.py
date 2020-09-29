import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import seaborn as sns
from matplotlib import animation, rc, rcParams
from IPython.display import HTML, Image
import random as rand

#----------------------------------------------------------------------------------------------------------------------------
#move table to very top right of excel sheet or program will fail


x_axis_col_num = 1 #defines the column numbers of the axes, change depending on excel sheet
y_axis_col_num = 2
additional_axis_col_num1 = 3
additional_axis_col_num2 = 4

x_axis_values = []
y_axis_values = []
additional_axis_values1 = []
additional_axis_values2 = []
additional_axis_values3 = []

#----------------------------------------------------------------------------------------------------------------------------

my_folder_path = "/Users/rossmcintyre/Documents/Python/Data/Matplotlib/"
file_name = "matplotlib_test.xlsx"
my_path = my_folder_path + file_name
my_wb_obj = openpyxl.load_workbook(my_path)

my_sheet_obj = my_wb_obj['Sheet1']

x_label = my_sheet_obj.cell(row = 1, column = x_axis_col_num).value
y_label = my_sheet_obj.cell(row = 1, column = y_axis_col_num).value
add1_label = my_sheet_obj.cell(row = 1, column = additional_axis_col_num1).value
add2_label = my_sheet_obj.cell(row = 1, column = additional_axis_col_num2).value

headings = []
for i in range(1, my_sheet_obj.max_column + 1):
    headings.append(my_sheet_obj.cell(row = 1, column = i).value)

column_data = []

for j in range(1, my_sheet_obj.max_column + 1):
    data = []
    for i in range(2, my_sheet_obj.max_row + 1):
        my_cell_obj = my_sheet_obj.cell(row = i, column = x_axis_col_num)
        my_cell_obj_value = my_sheet_obj.cell(row = i, column = y_axis_col_num)
        my_cell_obj_value1 = my_sheet_obj.cell(row = i, column = additional_axis_col_num1)
        my_cell_obj_value2 = my_sheet_obj.cell(row = i, column = additional_axis_col_num2)
        data_value = my_sheet_obj.cell(row = i, column = j)

        if my_cell_obj.value and j==1:
            data.append(str(data_value.value))

        if my_cell_obj.value and j!=1:
            data.append(data_value.value)

            

        if my_cell_obj.value and j==1:
            x_axis_values.append(str(my_cell_obj.value))
            y_axis_values.append(my_cell_obj_value.value)
            additional_axis_values1.append(my_cell_obj_value1.value)
            additional_axis_values2.append(my_cell_obj_value2.value)
            
    column_data.append(data)

#----------------------------------------------------------------------------------------------------------------------------

        
plot_type = input("Graph type? (Bar, Column, Line, Pie, Stacked Bar, Animated): ")
line_inputs = ['l', 'L', 'line', 'Line', 'line graph', 'Line Graph', 'Line graph', 'line Graph']
bar_inputs = ['b', 'B', 'bar', 'Bar', 'bar chart', 'Bar Chart', 'Bar chart', 'bar Chart']
column_inputs = ['c', 'C', 'column', 'Column', 'column chart', 'Column Chart', 'Column chart', 'column Chart']
pie_inputs = ['p', 'P', 'pie', 'Pie', 'pie chart', 'Pie Chart', 'Pie chart', 'pie Chart']
stacked_bar_inputs = ['sb', 'SB', 'Sb', 'sB', 'sbar', 'sBar', 'Sbar', 'SBar' 'stacked bar chart', 'Stacked Bar Chart', 'stacked Bar chart', 'Stacked bar Chart']
animate_inputs = ['a', 'ani', 'animate', 'animated', 'gif', 'Gif', 'GIF', 'animated graph', 'A', 'Ani', 'Animate', 'Animated', 'Animated graph' ]
#----------------------------------------------------------------------------------------------------------------------------

if plot_type in line_inputs:
    fig, ax = plt.subplots()
    ax.plot(x_axis_values, y_axis_values, 'bo:', markeredgecolor="g", markeredgewidth= 2, linewidth=0.5, markersize=12)
    ax.plot(x_axis_values, additional_axis_values1, 'yo:', markeredgecolor="r", markeredgewidth= 2, linewidth=1, markersize=10)
    ax.plot(x_axis_values, additional_axis_values2, 'go:', markeredgecolor="b", markeredgewidth= 2, linewidth=1.5, markersize=4)

    ax.set(xlabel=x_label, ylabel=y_label,
           title='Title')

    plt.xticks(rotation=60)
    ##plt.yticks(rotation=0)

    ax.grid(False)

    ##fig.savefig("test.png", transparent = True)
    plt.show()

#----------------------------------------------------------------------------------------------------------------------------

elif plot_type in bar_inputs:

    minvalue = 0
    maxvalue = 4
    number_of_ticks = 8
    
    x = np.arange(len(x_axis_values))
    for i in range(len(x)):
        x[i] = round(x[i], 2)
    width = 0.4  # the width of the bars
    

    fig, ax = plt.subplots()
    rects1 = ax.barh(x, y_axis_values, width, color = 'black') #add '-width/2 for dual column)
##    rects2 = ax.barh(x + width/2, additional_axis_values1, width, color = 'red')

    ax.set(xlabel=x_label, ylabel=y_label,
           title='Title')

    for i, v in enumerate(y_axis_values):
        ax.text(v + (maxvalue - minvalue)/100, i-0.18, str(round(v, 2)), color='black', fontweight='normal') #add '-width/2 to i for dual column)

##    for i, v in enumerate(additional_axis_values1):
##        ax.text(v + (maxvalue - minvalue)/100, i-0.18 + width/2, str(round(v, 2)), color='red', fontweight='normal')


    ax.set_yticks(x)
    ax.set_yticklabels(x_axis_values)

    plt.xticks(rotation=30)
    ##plt.yticks(rotation=0)

    ax.grid(False)

    ##fig.savefig("test.png", transparent = True)
    plt.show()

#----------------------------------------------------------------------------------------------------------------------------

elif plot_type in column_inputs:

    
    x = np.arange(len(x_axis_values))  # the label locations
    for i in range(len(x)):
        x[i] = round(x[i], 2)
    width = 0.4  # the width of the bars
    legend = []

    fig, ax = plt.subplots()
    rects1 = ax.bar(x - width/2, y_axis_values, width, color = 'black') #add '-width/2 for dual column)
    rects2 = ax.bar(x + width/2, additional_axis_values1, width, color = 'red')
##    rects3 = ax.bar(x + width/2, additional_axis_values2, width, color = 'red')

    ax.set(xlabel=x_label, ylabel=y_label,
           title='Title')

    def autolabel(rects):
##    """Attach a text label above each bar in *rects*, displaying its height."""
        for rect in rects:
            height = rect.get_height()
            ax.annotate('{}'.format(round(height, 2)),
                xy=(rect.get_x() + rect.get_width() / 2, height),
                xytext=(0, 3),  # 3 points vertical offset
                textcoords="offset points",
                ha='center', va='bottom')
    autolabel(rects1)
    autolabel(rects2)
##    autolabel(rects3)

    legend.append(rects1[0])
    legend.append(rects2[0])
##    legend.append(rects3[0])
    
    ax.set_xticks(x)
    ax.set_xticklabels(x_axis_values)
    plt.legend(legend, headings[1:3], loc = 'upper center', ncol = 1, fontsize = 8, fancybox = True,
               shadow = True)

    plt.xticks(rotation=64.83)
    ##plt.yticks(rotation=0)

    ax.grid(False)

    ##fig.savefig("test.png", transparent = True)
    plt.show()

#----------------------------------------------------------------------------------------------------------------------------

elif plot_type in pie_inputs:

    explode = [0]*len(x_axis_values)
    explode[3] = 0.5
    pie_labels = [str(x_axis_values[j]) + str('\n' + str(round(100*y_axis_values[j]/sum(y_axis_values), 2)) + '%') for j in range(0, len(x_axis_values))]

    sns.set_palette("Greys")
    
    fig, ax = plt.subplots()
    ax.pie(y_axis_values, explode= explode, labels= pie_labels, colors = None, autopct=None, pctdistance=0.6,
           shadow=False, labeldistance=1.1, startangle=None,
           radius=None, counterclock=True, wedgeprops=None, textprops=None,
           center=(0, 0), frame=False, rotatelabels=False, data=None)
    
    

    ax.set(title='Title')

    ax.grid(False)

    ##fig.savefig("test.png", transparent = True)
    plt.show()
        
#----------------------------------------------------------------------------------------------------------------------------
elif plot_type in stacked_bar_inputs:

    N = len(x_axis_values)
    ind = np.arange(N)
    width = 0.5
    legend = []
    bars_total = []
    sum_column = [0]*len(column_data[0])
    sum_column_total = []

    
    for n in range(1, len(column_data)-1):
        
        for jj in range(len(column_data[0])):

            sum_column[jj] += column_data[n][jj]
        
        if n==1:
            bars_total.append(plt.bar(ind, column_data[1], width, yerr=None))
            
        bars_total.append(plt.bar(ind, column_data[n+1], width,
             bottom = sum_column, yerr=None))
        
        
        if n==1:
            legend.append(bars_total[0][0])
            
        legend.append(bars_total[n][0])


        

    plt.ylabel('Weight (kg)')
    plt.title('Weight of food and omelettes and additional omelettes')
    plt.xticks(ind, column_data[0])
    plt.yticks(np.arange(0, 200, 10))
    plt.legend(legend, headings[1:], loc = 'upper center', ncol = 4, fontsize = 8, fancybox = True,
               shadow = True)

    ##fig.savefig("test.png", transparent = True)
    
    plt.show()
     

#----------------------------------------------------------------------------------------------------------------------------

if plot_type in animate_inputs:
    rcParams['animation.convert_path'] = r'/usr/local/Cellar/imagemagick/7.0.9-7/bin/convert'


    x = np.arange(0, 2*np.pi, 0.03)
    xx =np.arange(0, 6.5, 6.5/len(column_data[0]))
    y = np.sin(2.9*x)*np.cos(1.2*x)*np.sin(4.6*x)

    fig, ax = plt.subplots()
    line, = ax.plot(x, y, 'g', linewidth=1.5)
    sns.set_palette("Greys")

    def update(num, x, y, line):
        line.set_data(x[:num], y[:num])
        line.axes.axis([0, 6.5, -1.1, 1.1])
        return line,

    ani = animation.FuncAnimation(fig, update, len(x), fargs=[x, y, line],
                                  interval=10, blit=False, repeat = False)
    plt.xticks(xx, column_data[0])
    plt.show()
    # To save the animation, use e.g.
    #
    ani.save('/Users/rossmcintyre/Documents/Python/Outputs/Animations/animationtest.gif', writer='imagemagick', fps = 60)
    #
    # or
    #
    # from matplotlib.animation import FFMpegWriter
    # writer = FFMpegWriter(fps=15, metadata=dict(artist='Me'), bitrate=1800)
    # ani.save("movie.mp4", writer=writer)

    plt.show()
